# 請求書作成ボタンが押された際、すぐに実行するのではなくポップアップで確認を入れるようにした
# ファイルやディレクトリを指定しているかの確認を追加
# 関数化
# エラーハンドリングを追加
# メイン部分も関数化
# PDFファイル作成機能の追加（WindowsでExcelがインストールされていることが必要）
# Windows, macOS両方で動作するように処理を分ける（macOSではPDF作成機能を使わないように）
# version = "Ver.1.0.0"
from glob import glob
from datetime import datetime
from time import time
import openpyxl as opx
import pandas as pd
import PySimpleGUI as sg
import os
import platform

os_type = platform.system() # OSの種類を取得
if os_type == "Windows": # OSがWindowsだった場合、以下のモジュールをインポートする
    import win32com.client as win32
    import winreg # Windowsのレジストリの値を参照するため（Excelがインストールされているかどうかを確認する）

os.chdir(os.path.dirname(os.path.abspath(__file__))) # スクリプトの場所をカレントディレクトリにする
abs_dir = os.getcwd() # カレントディレクトリの絶対パスを取得
c_dir = os.path.basename(os.getcwd()) # カレントディレクトリのディレクトリ名のみ取得
excel_version = "15.0" # 必要なExcelのバージョンを指定。2016以上のバージョンは16.0

# ウィジェット部分の関数
def gui_widget(now_date):
    initial = "./"  # パスの初期値を定義
    
    # デフォルトで表示させておくファイルパスやフォルダパス
    default_template_file = f"{abs_dir}/invoice-template.xlsx" # GUIに表示しておくテンプレートファイルのデフォルトファイル名
    default_sales_folder = f"{abs_dir}/salesbooks" # GUIに表示しておく売上データフォルダのデフォルトフォルダ名
    default_save_folder = f"{abs_dir}/invoice" # GUIに表示しておく請求書保存先フォルダのデフォルトフォルダ名

    # ウィジェットの定義
    # WindowsとmacOSでウィジェットを変更する
    if os_type == "Windows": # Windowsの場合
        layout = [
        [sg.Text("テンプレートファイル選択"), sg.InputText(default_text=default_template_file, key="filepath"), sg.FileBrowse("Browse", initial_folder=initial, file_types=(("Excel files", "*.xlsx"), ("All files", "*.*")))],
        [sg.Text("売上データフォルダ選択", size=(21, 1)), sg.InputText(default_text=default_sales_folder, key="sales_folderpath", size=(45, 1)), sg.FolderBrowse("Browse", initial_folder=f'{initial}/salesbooks',)],
        [sg.Text("")],
        [sg.Text("請求書日付"), sg.Text("", size=(10, 1)), sg.InputText(default_text=now_date, key="date1", size=(46, 1)), sg.Button("本日", key="date")],
        [sg.Text("保存先フォルダ選択", size=(21, 1)), sg.InputText(default_text=default_save_folder, key="savefolder"), sg.FolderBrowse("Browse", initial_folder=f"{initial}/invoice")],
        [sg.Button("請求書を作成", key="gen", size=(10, 2)), sg.Text(key="result", size=(9, 1)), sg.Checkbox("PDFファイルも作成", key="plus_pdf")],
        [sg.Text("")],
        [sg.Button("終了", size=(10, 1))],
        ]
        return layout
    else: # macOSの場合
        layout = [
        [sg.Text("テンプレートファイル選択"), sg.InputText(default_text=default_template_file, key="filepath"), sg.FileBrowse("Browse", initial_folder=initial, file_types=(("Excel files", "*.xlsx"), ("All files", "*.*")))],
        [sg.Text("売上データフォルダ選択", size=(21, 1)), sg.InputText(default_text=default_sales_folder, key="sales_folderpath", size=(45, 1)), sg.FolderBrowse("Browse", initial_folder=f'{initial}/salesbooks',)],
        [sg.Text("")],
        [sg.Text("請求書日付"), sg.Text("", size=(10, 1)), sg.InputText(default_text=now_date, key="date1", size=(46, 1)), sg.Button("本日", key="date")],
        [sg.Text("保存先フォルダ選択", size=(21, 1)), sg.InputText(default_text=default_save_folder, key="savefolder"), sg.FolderBrowse("Browse", initial_folder=f"{initial}/invoice")],
        [sg.Button("請求書を作成", key="gen", size=(10, 2)), sg.Text(key="result", size=(9, 1))],
        [sg.Text("")],
        [sg.Button("終了", size=(10, 1))],
        ]
        return layout


# ファイルの存在確認を行う関数
def check_file_exists(value):
    # event（辞書）に入っている値を確認し、ファイルやディレクトリの存在確認を行う
    template_file_check = value.get("filepath") # PySimpleGUIで入力されたテンプレートファイルのパス
    sales_folder_check = value.get("sales_folderpath") # PySimpleGUIで入力された売上データディレクトリのパス
    save_folder_check = value.get("savefolder") # PySimpleGUIで入力された保存先ディレクトリのパス
    path_exists = os.path.isfile(template_file_check) and os.path.isdir(sales_folder_check) and os.path.isdir(save_folder_check) # ファイルとディレクトリが存在するかを確認
    return path_exists # True or False


# openpyxlでファイルを読み込み、pandasでdfにまとめる関数
def opx_to_pandas_df(filename_sales, window):
    window["result"].update("Excel作成中") # 実行中の動作表示
    window.refresh() # 表示更新
    df = pd.DataFrame(columns=["日付","購入者","品目", "個数", "値段", "小計"]) # データフレームの列を定義
    row = 0 # 行番号を定義。dfでは0から始まる
    for file in filename_sales: # xlsxファイルを順に変数fileに代入
        print(f"read_Excel_file: {file.replace("/", "\\")}") # 読み込みファイルを表示。最終的にはログファイルに記録したい
        wb_sales = opx.load_workbook(file, data_only=True) # xlsxファイルを数値で読み込む（Excelの計算式から計算結果を数値で書き込む）
        ws_sales = wb_sales["3月"] # 所定のワークシートを選択
        ws_title = ws_sales.title # シート名の取得（請求月表示に使用）
        max_row = ws_sales.max_row # シートの最終行を取得。openypxlで使う際は0始まりではなく1始まりなので+1する
        for r in range(4, max_row + 1):
            if ws_sales.cell(r, 1).value is not None: # セルが空でない場合
                df.loc[row, "日付"] = ws_sales.cell(r, 1).value # 日付をr行のA列から取得
                df.loc[row, "購入者"] = ws_sales.cell(r, 2).value # 購入者をr行のB列から取得
                df.loc[row, "品目"] = ws_sales.cell(r, 3).value # 品目をr行のC列から取得
                df.loc[row, "個数"] = ws_sales.cell(r, 4).value # 個数をr行のD列から取得
                df.loc[row, "値段"] = ws_sales.cell(r, 5).value # 値段をr行のE列から取得
                df.loc[row, "小計"] = ws_sales.cell(r, 6).value # 小計をr行のF列から取得
                row += 1 # 行を+1

        grouped = df.groupby("購入者") # 購入者毎にグループ化
    return grouped, ws_title # グループ化したデータとシート名を返す


#  データフレームを辞書に変換し、openpyxlでテンプレートファイルに書き込む関数
def write_to_excel(grouped, ws_title, template_file, cell_date, value):
    customers = {} # 購入者毎のデータを格納する辞書を作成
    excel_save_files = [] # 書き込んだファイル名を保存するリストを作成
    for name, group in grouped:
        customers[name] = group.reset_index(drop=True) # 購入者毎にグループ化した値を辞書に入れる

    for name, customer_df in customers.items():
        wb = opx.load_workbook(template_file) # 請求書テンプレートファイルを開く
        ws = wb.active # アクティブシートを取得

        for i, row in customer_df.iterrows(): # 購入者毎にループ
            ws["B4"] = row["購入者"] # 請求書テンプレートファイルの"B4"セルに購入者名を入力
            file_name = row["購入者"] # ファイル名
            without_space_file_name = file_name.replace(" ", "") # ファイル名からスペースを削除する
            ws["G3"] = cell_date # 請求書の日付
            ws["C10"] = f"{ws_title}分のご請求" # 件名
            bonding_value = f"{row['品目']}({row['日付'].strftime('%m/%d')})" # 品目と日付を1つのセルに書く(内訳欄)変数を定義
            ws.cell(row=i + 15, column=2, value=bonding_value) # 内訳欄
            ws.cell(row=i + 15, column=5, value=row["個数"]) # 個数欄
            ws.cell(row=i + 15, column=6, value=row["値段"]) # 単価欄
            ws.cell(row=i + 15, column=7, value=row["小計"]) # 金額(税込)欄
            # カーソルの位置などを調整する場合
            # ws.sheet_view.selection[0].activeCell = "A1"
            # ws.sheet_view.selection[0].sqref = "A1"
            # ws.sheet_view.topLeftCell = "A1"
            wb.properties.creator = "Automated Invoice Creation Project" # ファイルの作成者を任意に指定
            wb.properties.lastModifiedBy = "" # ファイルを前回保存者を任意に指定
        try:
            wb.save(f'{value["savefolder"]}/{without_space_file_name}様.xlsx') # 請求書の保存
            excel_save_files.append(f'{value["savefolder"]}/{without_space_file_name}様.xlsx') # 請求書の保存ファイル名をリストに追加（保存に成功したファイルをリストに追加。デバッグ用）
            print(f'write_Excel_file: {value["savefolder"].replace("\\", "/")}/{without_space_file_name}様.xlsx') # 保存ファイルを表示。最終的にはログファイルに記録したい
        except PermissionError as e:
            sg.popup_error(f"ファイルが他のアプリで開かれているようです。\n他のアプリを終了後、再度実行してください。\nエラー内容詳細: {e}", title="エラー") # ポップアップを表示
            print(excel_save_files)

            return False, excel_save_files
    return True, excel_save_files


# PDFファイルを作成する関数（WindowsでExcelがインストールされていることが必要）
def gen_pdf(value, abs_dir, window, excel_save_files):
    window["result"].update("PDF作成中") # 実行中の動作表示
    window.refresh() # 画面を更新
    save_folder = f'{value["savefolder"]}' # Excel保存先フォルダを取得
    file_names = [os.path.splitext(file)[0] for file in excel_save_files] # 拡張子を除いた保存先フォルダをリストで取得
    os.makedirs(f'{save_folder}/pdf', exist_ok=True) # 保存先フォルダにpdfというフォルダを作成。既にフォルダが存在した場合でも例外を発生させない
    excel = win32.Dispatch("Excel.Application") # pywin32を使ってExcelを起動させる
    excel.Visible = False # Excelのウィンドウを表示するかどうか。True:表示、False:非表示
    excel.DisplayAlerts = False # 警告メッセージを表示するかどうか。True:表示、False:非表示

    for file in file_names: # 保存先フォルダ内のxlsxファイルを順番に処理
        abs_dir, filename = os.path.split(file) # 絶対パスとファイル名に分ける
        save_pdf_file = f"{abs_dir.replace("/", "\\")}\\pdf\\{filename}.pdf" # PDF保存先のファイル名を代入。pywin32で操作をする場合、Windowsのディレクトリ区切り文字は\を使用する（/は使用できない模様）
        xlTypePDF = 0 # PDFを表す定数
        try:
            book = excel.Workbooks.Open(f"{file}.xlsx") # Excelファイルを開く
            print(f"read_Excel_file:{file.replace("\\", "/")}.xlsx") # 読み込みファイルを表示。最終的にはログファイルに記録したい
            book.ExportAsFixedFormat(xlTypePDF, save_pdf_file) # PDFファイルとして保存
            print(f"write_pdf_file:{save_pdf_file.replace("\\", "/")}") # 出力ファイルを表示。最終的にはログファイルに記録したい
        except Exception as e: # ディレクトリ区切りが"/"だった場合や保存先指定が不適切だった場合などの例外処理
            sg.popup_error(f"PDFファイルの保存時にエラーが発生しました。\n{e}", title="エラー") # ポップアップを表示
            window["result"].update("") # 実行中の動作表示
            window.refresh() # 画面を更新
            excel.Quit() # Excelを終了（Excelがバックグラウンドで実行したままになってしまうため）
            return False # 後処理で使うためFalseを返す
    excel.Quit() # Excelを終了
    return True # 正常に終了時Trueを返す


# Excelがインストールされているか、PDFファイルも作成にチェックが入っているか確認する関数（レジストリの値で確認）
def is_excel_installed(value):
    if os_type == "Windows": # OSがWindowsである場合のみ処理を行う
        if not value["plus_pdf"]: # "PDFファイルも作成"チェックが付いていない場合はTrueを返す（チェックを飛ばす）
            return True
        else:
            try:
                # レジストリキーのパス
                key_path = f"SOFTWARE\\Microsoft\\Office\\{excel_version}\\Excel\\InstallRoot" # Officeのレジストリパス
                with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, key_path) as key: # レジストリを開く
                    return True  # インストールされている
            except FileNotFoundError:
                sg.popup_error(f'指定されたバージョンのExcelがインストールされていません。\n"PDFファイルも作成" チェックを外して下さい。', title="エラー") # ポップアップを表示
                return False  # インストールされていない
    else: # OSがWindows以外の場合
        return True


# メイン関数
def main():
    today = datetime.now() # 日時（当日）を取得
    now_date = today.strftime("%Y年%m月%d日") # プログラム実行時に日付にする場合に使う

    window = sg.Window(f"請求書作成アプリ", gui_widget(now_date)) # ウィンドウを開く

    while True:
        event, value = window.read() # イベントと値を取得
        if event == sg.WIN_CLOSED: # ウィンドウの閉じるボタンが押された場合
            break

        # 終了ボタンが押された場合にポップアップ確認する動作
        if event == "終了":
            result = sg.popup_yes_no("アプリケーションを終了しますか？", title="確認")
            if result == "Yes":
                break

        # "本日"ボタンが押された場合に本日の日付に変える処理
        if event == "date":
            window["date1"].update(now_date) # 日付欄を更新する

        if event == "gen":
            if not is_excel_installed(value): # PDFを作成する場合、Excelがインストールされているかをチェック
                continue
            if check_file_exists(value):
                result = sg.popup_yes_no("請求書を作成しますか？", title="確認") # ポップアップを表示
                if result == "Yes":
                    window["result"].update("") # 動作中の表示を消す
                    window.refresh() # 画面更新
                    start_time = time() # 開始時間を記録
                    filename_sales = glob(f'{value["sales_folderpath"]}/*.xlsx') # 売上ファイル名をリストで取得
                    template_file = value["filepath"] # テンプレートファイルのパスを取得

                    cell_date = value["date1"] # 請求書日付エリアの文字を代入
                    
                    grouped, ws_title = opx_to_pandas_df(filename_sales,window) # 戻り値をそれぞれの変数に代入

                    #  データフレームを辞書に変換し、openpyxlでテンプレートファイルに書き込む
                    gen_result, excel_save_files = write_to_excel(grouped, ws_title, template_file, cell_date, value) # 戻り値はTrue / False 

                    if gen_result: # True, PDFを作成する場合でExcelがインストールされている場合
                        if os_type == "Windows" and value["plus_pdf"]:
                            if value["plus_pdf"]: # PDFを作成するにチェックが入っている場合
                                pdf_result = gen_pdf(value, abs_dir, window, excel_save_files) # PDFを生成する関数を呼び出す
                                if pdf_result: # True, 正常にPDFが作成された場合
                                    window["result"].update("完了")
                                    end_time = time() # 終了時間を記録
                                    est_time = round(end_time - start_time, 2) # 実行時間を計算
                                    print(f"{est_time}秒")
                                    sg.popup("Excelの請求書作成、PDFファイル作成が完了しました。", title="完了")

                        else: # False, PDFを作成しない場合
                            window["result"].update("完了")
                            end_time = time() # 終了時間を記録
                            est_time = round(end_time - start_time, 2) # 実行時間を計算
                            print(f"{est_time}秒")
                            sg.popup("請求書作成が完了しました。", title="完了") # ポップアップを表示

            else:
                sg.popup("ファイルやフォルダの指定に不備があります。\n確認してください。", title="確認") # ポップアップを表示

    window.close() # ウィンドウを閉じる


# メイン関数を実行
if __name__ == "__main__":
    main()
